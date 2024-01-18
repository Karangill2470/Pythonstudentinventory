import openpyxl
from datetime import datetime
import re

# Verification function for student ID format
def validate_id_of_Student (input_id):
    # Alphanumeric characters are matched by a regular expression pattern
    pattern = r'^[A-Za-z0-9]+$'
    return bool(re.match(pattern, input_id))

# Validation function for semester date format
def validate_dates_of_Semester(input_date):
    try:
        datetime.strptime(input_date, "%m/%d/%Y")
    except ValueError:
        try:
            datetime.strptime(input_date, "%d-%m-%Y")
        except ValueError:
            return False
    return True

#Function to verify the student's CGPA format
def validate_CGPA_of_Student(input_cgpa):
    if input_cgpa.upper() == 'N/A':
        return True
    pattern = r'^\d+(\.\d{1,2})?$'
    return bool(re.match(pattern, input_cgpa))

# Function to determine a student's row number by name
def find_studentname(sheet, name):
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == name:
            return i
    return None

# Capability to show every student record
def display_students_records(sheet):
    print('\nCurrent Student Records:\n')
    print(f'{"Name":<20}{"ID":<10}{"Start Date":<15}{"End Date":<15}{"CGPA":<10}{"Course":<20}{"Duration":<15}')
    for i in range(2, sheet.max_row + 1):
        # Get values from the current row's cells.
        name = sheet.cell(row=i, column=1).value
        student_id = sheet.cell(row=i, column=2).value
        start_date = sheet.cell(row=i, column=3).value
        end_date = sheet.cell(row=i, column=4).value
        cgpa = sheet.cell(row=i, column=5).value
        course = sheet.cell(row=i, column=6).value
        # Use the calculate_semester_duration function to determine the duration.
        duration = calculate_semester_duration(start_date, end_date)
        # Publish information from student records in format.
        print(f'{name:<20}{student_id:<10}{start_date:<15}{end_date:<15}{cgpa:<10}{course:<20}{duration:<15}')

# Function to determine the length of a semester
def calculate_semester_duration(start_date, end_date):
    if start_date is None or end_date is None:
        raise ValueError("Start date and end date cannot be None.")
    
    # Experiment with parsing the start_date in various date formats.
    try:
        start_date_obj = datetime.strptime(start_date, "%m/%d/%Y")
    except ValueError:
        try:
            start_date_obj = datetime.strptime(start_date, "%d-%m-%Y")
        except ValueError:
            raise ValueError("Invalid start date format. Please use MM/DD/YYYY or DD-MM-YYYY.")
        
    # Experiment with parsing the end_date in various date formats.
    try:
        end_date_obj = datetime.strptime(end_date, "%m/%d/%Y")
    except ValueError:
        try:
            end_date_obj = datetime.strptime(end_date, "%d-%m-%Y")
        except ValueError:
            raise ValueError("Invalid end date format. Please use MM/DD/YYYY or DD-MM-YYYY.")

    duration = end_date_obj - start_date_obj
    return f'{duration.days} days'

# Password authentication function
def Password_function():
    correct_password = "7070"
    attempts_left = 4

    while attempts_left > 0:
        user_password = input("Enter password: ")
        if user_password == correct_password:
            return True
        else:
            attempts_left -= 1
            print(f'Incorrect password. Attempts left: {attempts_left}')

    print("Authentication failed. Exiting program.")
    return False

def main():
    # File path for the Excel workbook
    workbook_path = r'D:\Semester 2\Document Automation Python\Final Project\student_Record_File.xlsx'
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.active

    if not Password_function():
        exit()

    print('\n=== Welcome to the Custom Student Records System ===')
    while True:
        print('\nOptions:')
        print('1. Add a new student record')
        print('2. Update an existing student record')
        print('3. Delete a student record')
        print('4. View the list of student records')
        print('5. Search and display a specific student record')
        print('6. Exit')

        choice_of_user= input('Enter your choice you want to do: ')
        # Option to add a new student record
        if choice_of_user== '1':
            student_name = input('Enter the name of the student: ')
            row = find_studentname(worksheet, student_name)
            if row is not None:
                print('Student record already exists.')
            else:
                #Ask the user for additional information if the student record is missing.
                student_id = input('Enter the student ID: ')
                while not validate_id_of_Student (student_id):
                    student_id = input(int('Invalid input. Enter the student ID: '))
                start_date = input('Enter the start date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                while not validate_dates_of_Semester (start_date):
                    start_date = input('Invalid input. Enter the start date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                end_date = input('Enter the end date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                while not validate_dates_of_Semester (end_date):
                    end_date = input('Invalid input. Enter the end date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                input_cgpa = input('Enter the CGPA of the student (or N/A if not applicable): ')
                while not validate_CGPA_of_Student(input_cgpa):
                    input_cgpa = input('Invalid input. Enter the CGPA of the student (or N/A if not applicable): ')
                cgpa = 'N/A' if input_cgpa.upper() == 'N/A' else f'{float(input_cgpa):.2f}%'
                course = input('Enter the course of the student: ').upper()

                worksheet.append([student_name, student_id, start_date, end_date, cgpa, course])
                workbook.save(workbook_path)
                print('Student record added successfully!')

        # Option to update an existing student record
        elif choice_of_user== '2':
            student_name = input('Enter the name of the student to update: ')
            row = find_studentname(worksheet, student_name)

            if row is not None:
                # Show the specifics of the discovered student record.
                print(f"\nStudent Record Found:\nName: {worksheet.cell(row=row, column=1).value}\nID: {worksheet.cell(row=row, column=2).value}\nStart Date: {worksheet.cell(row=row, column=3).value}\nEnd Date: {worksheet.cell(row=row, column=4).value}\nCGPA: {worksheet.cell(row=row, column=5).value}\nCourse: {worksheet.cell(row=row, column=6).value}")

                new_id = input('Enter the new student ID: ')
                while not validate_id_of_Student (new_id):
                    new_id = input('Invalid input. Enter the new student ID: ')

                new_start_date = input('Enter the new start date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                while not validate_dates_of_Semester(new_start_date):
                    new_start_date = input('Invalid input. Enter the new start date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')

                new_end_date = input('Enter the new end date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')
                while not validate_dates_of_Semester(new_end_date):
                    new_end_date = input('Invalid input. Enter the new end date of the semester (MM/DD/YYYY or DD-MM-YYYY): ')

                new_input_cgpa = input('Enter the new CGPA of the student (or N/A if not applicable): ')
                while not validate_CGPA_of_Student(new_input_cgpa):
                    new_input_cgpa = input('Invalid input. Enter the new CGPA of the student (or N/A if not applicable): ')
                new_cgpa = 'N/A' if new_input_cgpa.upper() == 'N/A' else f'{float(new_input_cgpa):.2f}%'

                new_course = input('Enter the new course of the student: ').upper()
                # Add the new information for the discovered student record to the worksheet.
                worksheet.cell(row=row, column=2, value=new_id)
                worksheet.cell(row=row, column=3, value=new_start_date)
                worksheet.cell(row=row, column=4, value=new_end_date)
                worksheet.cell(row=row, column=5, value=new_cgpa)
                worksheet.cell(row=row, column=6, value=new_course)

                workbook.save(workbook_path)
                print('Student record updated successfully!')

            else:
                print('Student record not found.')
                
        # Option to delete a student record
        elif choice_of_user== '3':
            student_name = input('Enter the name of the student to delete: ')
            row = find_studentname(worksheet, student_name)

            if row is not None:
                worksheet.delete_rows(row, amount=1)
                workbook.save(workbook_path)
                print('Student record deleted successfully!')

            else:
                print('Student record not found.')
                
        # Option to view the list of student records
        elif choice_of_user== '4':
            display_students_records (worksheet)

        # Option to search and display a specific student record
        elif choice_of_user== '5':
            print('\nSearch and display a specific student record: ')
            student_name = input('Enter the name of the student: ')
            row = find_studentname(worksheet, student_name)

            if row is not None:
                print(f'\nStudent Record Found:\nName: {worksheet.cell(row=row, column=1).value}\nID: {worksheet.cell(row=row, column=2).value}\nStart Date: {worksheet.cell(row=row, column=3).value}\nEnd Date: {worksheet.cell(row=row, column=4).value}\nCGPA: {worksheet.cell(row=row, column=5).value}\nCourse: {worksheet.cell(row=row, column=6).value}')
            else:
                print('Student record not found.')
                
        # Option to exit the program
        elif choice_of_user== '6':
            print('\nExiting program of Final Project...')
            break

        else:
            print('\nInvalid choice. Please try again.')

if __name__ == "__main__":
    main()
